from flask import Flask, request, jsonify, render_template
import openpyxl
from openpyxl import load_workbook
import os

app = Flask(__name__)
EXCEL_AUTORES = "C:/Users/andres.guerra.d/uniminuto.edu/G-SUBDIRECCIÓN CENTRO EDITORIAL - Matrices/Autores publicaciones.xlsx"
EXCEL_PUBLICACIONES = "C:/Users/andres.guerra.d/uniminuto.edu/G-SUBDIRECCIÓN CENTRO EDITORIAL - Matrices/Publicaciones_CE.xlsx"
EXCEL_MATRIZ = "C:/Users/andres.guerra.d/uniminuto.edu/G-SUBDIRECCIÓN CENTRO EDITORIAL - Matrices/Matriz Intermedia.xlsx"


@app.route("/")
def home():
    return render_template("home.html")

@app.route("/autor", methods=["GET"])
def formulario_autor():
    return render_template("autores.html")

@app.route("/publicacion", methods=["GET"])
def formulario_publicacion():
    return render_template("agregar_publicacion.html")

@app.route("/matriz-intermedia", methods=["GET"])
def formulario_matriz():
    return render_template("formulario_matriz.html")

@app.route("/buscar", methods=["GET"])
def buscar():
    valor_input = request.args.get("q", "").strip().lower()
    coincidencias = []

    if not os.path.exists(EXCEL_AUTORES):
        return jsonify([])

    wb = load_workbook(EXCEL_AUTORES)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        documento = str(row[0]).strip().lower() if row[0] else ""
        nombre = str(row[1]).strip().lower() if row[1] else ""
        if valor_input in documento or valor_input in nombre:
            coincidencias.append(f"{row[0]} - {row[1]}")

    return jsonify(coincidencias)

@app.route("/guardar", methods=["POST"])
def guardar():
    from openpyxl.utils import get_column_letter

    data = request.json
    nombre = data.get("nombre", "").strip().lower()

    if not os.path.exists(EXCEL_AUTORES):
        return jsonify({"status": "error", "mensaje": "Archivo Excel no encontrado"}), 404

    wb = load_workbook(EXCEL_AUTORES)
    ws = wb.active

    # Verificar duplicado por nombre
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] and str(row[1]).strip().lower() == nombre:
            return jsonify({"status": "existe", "mensaje": "Este autor ya está registrado"})

    # Agregar nueva fila
    nueva_fila = [
        data.get("documento"), data.get("nombre"), data.get("pseudonimo"),
        data.get("sexo"), data.get("perfil"), data.get("nacionalidad"),
        data.get("correo"), data.get("formacion"), "",
        data.get("rectoria_original"), data.get("centro"), data.get("facultad"),
        data.get("programa"), data.get("filiacion"), data.get("pais"),
        data.get("investigador"), data.get("fecha")
    ]
    ws.append(nueva_fila)

    # Actualizar rango de la tabla "Autores_publicaciones"
    tabla = ws.tables["Autores_publicaciones"]
    ref = tabla.ref  # Ej: "A1:U3428"
    inicio, fin = ref.split(":")
    nueva_fila_num = ws.max_row
    col_inicio = inicio[:1]
    col_fin = get_column_letter(ws.max_column)
    tabla.ref = f"{inicio}:{col_fin}{nueva_fila_num}"

    wb.save(EXCEL_AUTORES)
    return jsonify({"status": "ok", "mensaje": "Autor agregado correctamente"})

@app.route("/guardar-publicacion", methods=["POST"])
def guardar_publicacion():
    try:
        data = request.json

        if not os.path.exists(EXCEL_PUBLICACIONES):
            return jsonify({"status": "error", "mensaje": "Archivo Excel no encontrado"}), 404

        wb = load_workbook(EXCEL_PUBLICACIONES)
        ws = wb.active
        tabla = ws.tables["Publicaciones_CE"]
        ref = tabla.ref

        isbn = (data.get("isbn") or "").strip().lower()
        isbn2 = (data.get("isbn2") or "").strip().lower()
        eisbn = (data.get("eisbn") or "").strip().lower()
        eisbn2 = (data.get("eisbn2") or "").strip().lower()
        codigo = (data.get("codigo_publicacion") or "").strip().lower()

        for row in ws.iter_rows(min_row=2, values_only=True):
            fila = [str(cell).strip().lower() if cell else "" for cell in row]
            if ((isbn and isbn in fila[:8]) or (isbn2 and isbn2 in fila[:8]) or
                (eisbn and eisbn in fila[:8]) or (eisbn2 and eisbn2 in fila[:8]) or
                (codigo and codigo == fila[0])):
                return jsonify({"status": "duplicado"})

        nueva_fila = [
            data.get("codigo_publicacion"), data.get("titulo_libro"), data.get("subtitulo"),
            "", data.get("isbn"), data.get("isbn2"), data.get("eisbn"),
            data.get("eisbn2"), data.get("doi_libro"), data.get("doi_capitulo"),
            data.get("registro_dinda"), data.get("titulo_capitulo"), data.get("anio"),
            data.get("mes"), data.get("fecha"), data.get("nombre_proyecto"),
            data.get("codigo_proyecto"), data.get("linea_editorial"), data.get("tipologia"),
            data.get("thema"), data.get("area_conocimiento"), data.get("materia"),
            data.get("palabras_clave"), data.get("nombre_coleccion"),
            data.get("isbn_issn_coleccion"), data.get("formato"), data.get("url_repositorio"),
            data.get("idioma"), data.get("resumen"), data.get("ods"),
            data.get("rectoria_normalizada"), data.get("rectoria_original"),
            data.get("centro_universitario"), data.get("origen"), data.get("financiador"),
            data.get("fecha_diligenciamiento")
        ]
        ws.append(nueva_fila)

        # Expandir la tabla
        from openpyxl.utils import range_boundaries, get_column_letter
        min_col, min_row, max_col, max_row = range_boundaries(tabla.ref)
        nueva_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{ws.max_row}"
        tabla.ref = nueva_ref

        wb.save(EXCEL_PUBLICACIONES)
        return jsonify({"status": "ok"})

    except Exception as e:
        print("⚠️ ERROR EN EL BACKEND:", e)
        return jsonify({"status": "error", "mensaje": str(e)}), 500

@app.route("/guardar-matriz", methods=["POST"])
def guardar_matriz():
    try:
        data = request.json
        registros = data.get("registros", [])

        if not os.path.exists(EXCEL_MATRIZ):
            return jsonify({"status": "error", "mensaje": "Archivo Excel no encontrado"}), 404

        wb = load_workbook(EXCEL_MATRIZ)
        ws = wb.active
        tabla = ws.tables["Tabla2"]
        ref = tabla.ref

        for reg in registros:
            fila = [
                reg.get("codigo"),
                reg.get("titulo"),
                reg.get("nombre"),
                reg.get("nombre"),
                reg.get("documento")
            ]
            ws.append(fila)

        # Expandir la tabla
        from openpyxl.utils import range_boundaries, get_column_letter
        min_col, min_row, max_col, max_row = range_boundaries(tabla.ref)
        nueva_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{ws.max_row}"
        tabla.ref = nueva_ref

        wb.save(EXCEL_MATRIZ)
        return jsonify({"status": "ok"})

    except Exception as e:
        print("⚠️ ERROR GUARDANDO MATRIZ:", e)
        return jsonify({"status": "error", "mensaje": str(e)}), 500

@app.route("/api/publicaciones")
def api_publicaciones():
    if not os.path.exists(EXCEL_PUBLICACIONES):
        return jsonify([])
    wb = load_workbook(EXCEL_PUBLICACIONES)
    ws = wb.active
    publicaciones = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        codigo = str(row[0]).strip()
        titulo = str(row[1]).strip()
        if codigo and titulo:
            publicaciones.append({"codigo": codigo, "titulo": titulo})
    return jsonify(publicaciones)

@app.route("/api/autores")
def api_autores():
    if not os.path.exists(EXCEL_AUTORES):
        return jsonify([])
    wb = load_workbook(EXCEL_AUTORES)
    ws = wb.active
    autores = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        documento = str(row[0]).strip()
        nombre = str(row[1]).strip()
        if documento and nombre:
            autores.append({"documento": documento, "nombre": nombre})
    return jsonify(autores)

if __name__ == "__main__":
    app.run(debug=True)

